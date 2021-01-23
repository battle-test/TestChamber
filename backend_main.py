import sys
from PyQt5 import QtWidgets
from frontend_main import Ui_MainWindow
import pandas as pd
import numpy as np
import openpyxl as opx
import os

# Create app
app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(MainWindow)
MainWindow.show()


class Protocol:
    """Processes the value of the physical parameters obtained during measurement (temperature and moisture),
    and then saves it to the finished Excel file."""

    def __init__(self):
        self.data_set = ""
        self.type_sep_field = ui.frame_openFile_comboBox_typeSepField.setCurrentIndex(0)
        self.device = ui.frame_openFile_comboBox_device.setCurrentIndex(0)
        self.type_encoding = ui.frame_openFile_comboBox_typeEncoding.setCurrentIndex(0)
        self.type_decimal = ui.frame_openFile_comboBox_typeDecimal.setCurrentIndex(0)
        self.qty_sensors = ui.frame_openFile_comboBox_qtySensors.setCurrentIndex(0)
        self.list_temp = [
            ui.frame_temp_lineEdit_temp1.setText(''),
            ui.frame_temp_lineEdit_temp2.setText(''),
            ui.frame_temp_lineEdit_temp3.setText(''),
            ui.frame_temp_lineEdit_temp4.setText(''),
            ui.frame_temp_lineEdit_temp5.setText(''),
            ui.frame_temp_lineEdit_temp6.setText(''),
            ui.frame_temp_lineEdit_temp7.setText(''),
            ui.frame_temp_lineEdit_temp8.setText('')
        ]
        self.list_hum = [
            ui.frame_hum_lineEdit_hum1.setText(''),
            ui.frame_hum_lineEdit_hum2.setText(''),
            ui.frame_hum_lineEdit_hum3.setText(''),
            ui.frame_hum_lineEdit_hum4.setText(''),
            ui.frame_hum_lineEdit_hum5.setText(''),
        ]
        self.list_temp_in_hum = [
            ui.frame_hum_lineEdit_temp1.setText(''),
            ui.frame_hum_lineEdit_temp2.setText(''),
            ui.frame_hum_lineEdit_temp3.setText(''),
            ui.frame_hum_lineEdit_temp4.setText(''),
            ui.frame_hum_lineEdit_temp5.setText(''),
        ]
        self.excel_pattern = os.getcwd() + "\\source9.xlsm"
        self.excel_delete_empty_hum = ui.frame_optionFile_checkBox_deleteSheetTemp.setChecked(True)
        self.excel_delete_empty_temp_hum = ui.frame_optionFile_checkBox_deleteSheetHum.setChecked(True)
        self.data_clean_temp = np.array
        self.data_clean_hum = np.array
        self.dict_temp = {}
        self.dict_hum = {}
        self.dict_temp_in_hum = {}
        self.excel_file_path = 'C:\\'
        self.excel_file = ''

    def open_data_set(self):
        ui.frame_temp_lineEdit_temp1.setText('-20')
        ui.frame_temp_lineEdit_temp2.setText('10')
        ui.frame_temp_lineEdit_temp3.setText('40')
        ui.frame_temp_lineEdit_temp4.setText('70')
        ui.frame_temp_lineEdit_temp5.setText('100')
        ui.frame_hum_lineEdit_hum1.setText('20')
        ui.frame_hum_lineEdit_temp1.setText('85')
        ui.frame_hum_lineEdit_hum2.setText('95')
        ui.frame_hum_lineEdit_temp2.setText('15')

        self.data_set = ''
        data_set = QtWidgets.QFileDialog.getOpenFileNames(None, "Open FIle", "",
                                                          "XLS, CSV, TXT file (*.xls; *.csv; *.txt)")
        data_set = data_set[0]
        if len(data_set) > 1:
            ui.frame_openFile_lineEdit.setText(f"Количество загруженных файлов: {len(data_set)}")
        elif len(data_set) == 1:
            ui.frame_openFile_lineEdit.setText(f"{data_set[0]}")
        else:
            pass
        self.data_set = data_set

    def open_excel_pattern(self):
        excel_pattern = QtWidgets.QFileDialog.getOpenFileNames(None, "Open FIle", "",
                                                               "XLSM file (*.xlsm)")
        try:
            excel_pattern = excel_pattern[0][0]
            self.excel_pattern = excel_pattern
        except IndexError:
            pass

    def save_protocol(self):
        self.type_sep_field = ui.frame_openFile_comboBox_typeSepField.currentText()
        self.device = ui.frame_openFile_comboBox_device.currentText()
        self.type_encoding = ui.frame_openFile_comboBox_typeEncoding.currentText()
        self.type_decimal = ui.frame_openFile_comboBox_typeDecimal.currentText()
        self.qty_sensors = int(ui.frame_openFile_comboBox_qtySensors.currentText())
        self.list_temp = filter(None, [
            ui.frame_temp_lineEdit_temp1.text(),
            ui.frame_temp_lineEdit_temp2.text(),
            ui.frame_temp_lineEdit_temp3.text(),
            ui.frame_temp_lineEdit_temp4.text(),
            ui.frame_temp_lineEdit_temp5.text(),
            ui.frame_temp_lineEdit_temp6.text(),
            ui.frame_temp_lineEdit_temp7.text(),
            ui.frame_temp_lineEdit_temp8.text(),
        ])
        self.list_hum = filter(None, [
            ui.frame_hum_lineEdit_hum1.text(),
            ui.frame_hum_lineEdit_hum2.text(),
            ui.frame_hum_lineEdit_hum3.text(),
            ui.frame_hum_lineEdit_hum4.text(),
            ui.frame_hum_lineEdit_hum5.text(),
        ])
        self.list_temp_in_hum = filter(None, [
            ui.frame_hum_lineEdit_temp1.text(),
            ui.frame_hum_lineEdit_temp2.text(),
            ui.frame_hum_lineEdit_temp3.text(),
            ui.frame_hum_lineEdit_temp4.text(),
            ui.frame_hum_lineEdit_temp5.text(),
        ])
        self.excel_pattern = 'Path file xls'
        self.excel_delete_empty_hum = ui.frame_optionFile_checkBox_deleteSheetTemp.isChecked()
        self.excel_delete_empty_temp_hum = ui.frame_optionFile_checkBox_deleteSheetHum.isChecked()

        def parser_file():
            def parser_mit_file():
                if len(self.data_set) == 1:
                    self.data_set = self.data_set[0]
                    df = pd.read_csv(self.data_set, sep=self.type_sep_field, encoding=self.type_encoding,
                                     header=None, skiprows=5, na_values='#', decimal=self.type_decimal)
                    df = df.iloc[:, 1:self.qty_sensors + 2]
                    df = df.dropna()
                    # df_time = df.iloc[:, 0]
                    df_temp = df.iloc[:, 1:self.qty_sensors + 1]
                    self.data_clean_temp = np.array(df_temp)
                elif len(self.data_set) > 1:
                    df = pd.concat(
                        [pd.read_csv(f, sep=self.type_sep_field, encoding=self.type_encoding, engine='python',
                                     header=None, skiprows=5, na_values='#', decimal=self.type_decimal) for f
                         in self.data_set], axis=0, ignore_index=True)
                    df = df.iloc[:, 1:self.qty_sensors + 2]
                    df = df.dropna()
                    # df_time = df.iloc[:, 0]
                    df_temp = df.iloc[:, 1:self.qty_sensors + 1]
                    self.data_clean_temp = np.array(df_temp)

            def parser_rotronic_file():
                df = pd.concat(
                    [pd.read_csv(f, usecols=[2, 3], sep=self.type_sep_field, encoding=self.type_encoding,
                                 engine='python',
                                 header=None, skiprows=57, na_values=' --.--', decimal=self.type_decimal) for f in
                     self.data_set], axis=1, ignore_index=True)
                # df_time = pd.concat(
                #    [pd.read_csv(f, usecols=[1], sep=self.type_sep_field, encoding=self.type_encoding, engine='python',
                #                  header=None, skiprows=57, na_values=' --.--', decimal=self.type_decimal) for f in
                #                   self.data_set], axis=1, ignore_index=True)
                df_hum = df.iloc[:, ::2]
                df_hum = df_hum.dropna()
                df_temp = df.iloc[:, 1::2]
                df_temp = df_temp.dropna()
                # df_time = df_time.iloc[:len(df_temp), 0:1]
                self.data_clean_temp = np.array(df_temp)
                self.data_clean_hum = np.array(df_hum)

            if ui.frame_openFile_comboBox_device.currentIndex() == 0:
                parser_mit_file()

            elif ui.frame_openFile_comboBox_device.currentIndex() == 1:
                parser_rotronic_file()

        parser_file()

        def get_temp_result():
            self.list_temp = list(map(int, filter(None, self.list_temp)))
            array_abs = np.array([[0, 0, 0],
                                  [0, 0, 0]])
            self.dict_temp = dict.fromkeys(self.list_temp, array_abs)
            for temp in self.dict_temp:
                for idx, row in enumerate(self.data_clean_temp):
                    if idx < (len(self.data_clean_temp) - 1):
                        if temp - 2 < np.mean(row) < temp + 2:
                            abs = np.sum(np.abs(np.absolute(self.data_clean_temp[idx]) -
                                                np.absolute(self.data_clean_temp[idx + 1])))
                            if abs < 0.5:
                                self.dict_temp[temp] = np.append(self.dict_temp[temp],
                                                                 [[idx, abs, (np.mean(self.data_clean_temp[idx]))]],
                                                                 axis=0)
                self.dict_temp[temp] = self.dict_temp[temp][2:]
            for temp in self.dict_temp:
                l = range(len(self.data_clean_temp))
                best_set = set(list(self.dict_temp[temp][:, 0]))
                k = 30
                for i in range(len(l) - k + 1):
                    if best_set.issuperset(l[i:i + k]):
                        best_result = i
                        self.dict_temp[temp] = best_result
            for temp in self.dict_temp:
                self.dict_temp[temp] = ((pd.DataFrame(self.data_clean_temp)).
                                            iloc[self.dict_temp[temp]:self.dict_temp[temp] + 30, 0:9])

        get_temp_result()

        def get_hum_result():
            self.list_temp_in_hum = list(map(int, filter(None, self.list_temp_in_hum)))
            self.list_hum = list(map(int, filter(None, self.list_hum)))
            array_abs = np.array([[0, 0, 0],
                                  [0, 0, 0]])
            self.dict_hum = dict.fromkeys(self.list_hum, array_abs)
            self.dict_temp_in_hum = dict.fromkeys(self.list_temp_in_hum, array_abs)
            for hum in self.dict_hum:
                for idx, row in enumerate(self.data_clean_hum):
                    if idx < (len(self.data_clean_hum) - 1):
                        if hum - 2 < np.mean(row) < hum + 2:
                            abs = np.sum(np.abs(np.absolute(self.data_clean_hum[idx]) -
                                                np.absolute(self.data_clean_hum[idx + 1])))
                            if abs < 3:
                                self.dict_hum[hum] = np.append(self.dict_hum[hum],
                                                               [[idx, abs, (np.mean(self.data_clean_hum[idx]))]],
                                                               axis=0)
                self.dict_hum[hum] = self.dict_hum[hum][2:]
            for hum in self.dict_hum:
                l = range(len(self.data_clean_hum))
                best_set = set(list(self.dict_hum[hum][:, 0]))
                k = 30
                for i in range(len(l) - k + 1):
                    if best_set.issuperset(l[i:i + k]):
                        best_result = i
                        self.dict_hum[hum] = best_result
            for hum in self.dict_hum:
                key = self.list_temp_in_hum[self.list_hum.index(hum)]
                self.dict_temp_in_hum[key] = ((pd.DataFrame(self.data_clean_temp)).
                                                  iloc[self.dict_hum[hum]:self.dict_hum[hum] + 30, 0:9])
                self.dict_hum[hum] = ((pd.DataFrame(self.data_clean_hum)).
                                          iloc[self.dict_hum[hum]:self.dict_hum[hum] + 30, 0:9])

        get_hum_result()

        def save_result():
            temp_name_xl_lists = [f't{x}' for x in range(1, 9)]
            hum_name_xl_lists = [f'rh_t{x}' for x in range(1, 6)]
            delete_temp_name_xl_lists = temp_name_xl_lists[len(self.list_temp):]
            delete_hum_name_xl_lists = hum_name_xl_lists[len(self.list_hum):]
            temp_name_xl_lists = temp_name_xl_lists[:len(self.list_temp)]
            hum_name_xl_lists = hum_name_xl_lists[:len(self.list_hum)]
            self.excel_pattern = "source9.xlsm"
            wb = opx.load_workbook("source9.xlsm", keep_vba=True)

            def save_temp():
                for delete_temp_name_xl_list in delete_temp_name_xl_lists:
                    wb.remove(wb.get_sheet_by_name(str(delete_temp_name_xl_list)))
                for temp_name_xl_list in temp_name_xl_lists:
                    xl_iterator = wb[(str(temp_name_xl_list))].iter_rows(min_col=3, max_col=11, min_row=3, max_row=32)
                    wb[(str(temp_name_xl_list))].cell(row=39, column=3).value = self.list_temp[0]
                    data = self.dict_temp.pop(self.list_temp.pop(0))
                    for i in range(30):
                        rows_excel = list(next(xl_iterator))
                        rows_data = data.iloc[i]
                        zip(rows_excel, rows_data)
                        for row, row_data in zip(rows_excel, rows_data):
                            row.value = row_data

            def save_hum():
                for delete_hum_name_xl_list in delete_hum_name_xl_lists:
                    wb.remove(wb.get_sheet_by_name(str(delete_hum_name_xl_list)))
                for hum_name_xl_list in hum_name_xl_lists:
                    xl_iterator = wb[(str(hum_name_xl_list))].iter_rows(min_col=3, max_col=11, min_row=3, max_row=32)
                    wb[(str(hum_name_xl_list))].cell(row=39, column=3).value = self.list_hum[0]
                    data_hum = self.dict_hum.pop(self.list_hum.pop(0))
                    for i in range(30):
                        rows_excel = list(next(xl_iterator))
                        rows_data = data_hum.iloc[i]
                        zip(rows_excel, rows_data)
                        for row, row_data in zip(rows_excel, rows_data):
                            row.value = row_data
                    xl_iterator = wb[(str(hum_name_xl_list))].iter_rows(min_col=3, max_col=11, min_row=50, max_row=79)
                    wb[(str(hum_name_xl_list))].cell(row=86, column=3).value = self.list_temp_in_hum[0]
                    data_temp = self.dict_temp_in_hum.pop(self.list_temp_in_hum.pop(0))
                    for i in range(30):
                        rows_excel = list(next(xl_iterator))
                        rows_data = data_temp.iloc[i]
                        zip(rows_excel, rows_data)
                        for row, row_data in zip(rows_excel, rows_data):
                            row.value = row_data

            save_temp()
            save_hum()
            self.excel_file = wb
            self.excel_file_path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Save File", "my_protocol",
                                                                            "xlsm file, (*.xlsm)")
            self.excel_file.save(self.excel_file_path)

        save_result()


PROTOCOL = Protocol()
ui.frame_openFile_pushButton.clicked.connect(PROTOCOL.open_data_set)
ui.frame_optionFile_pushButton_modForm.clicked.connect(PROTOCOL.open_excel_pattern)
ui.frame_optionFile_saveProtocol.clicked.connect(PROTOCOL.save_protocol)
ui.frame_optionFile_reset.clicked.connect(PROTOCOL.__init__)

sys.exit(app.exec_())
# было дело...
